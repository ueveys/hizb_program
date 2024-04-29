from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import random
from openpyxl.styles import PatternFill
import random

def calculate_start_points(people_names):
    num_people = len(people_names)
    num_cuz = 30
    num_hizb_per_cuz = 4
    start_points = []

    current_cuz = 1
    current_hizb = 1
    for _ in range(num_people):
        start_points.append((current_cuz, current_hizb))
        current_hizb += 1
        if current_hizb > num_hizb_per_cuz:
            current_hizb = 1
            current_cuz += 1
            if current_cuz > num_cuz:
                current_cuz = 1

    return start_points

def generate_schedule(start_date, end_date, people_names):
    start_points = calculate_start_points(people_names)
    num_cuz = 30
    num_hizb_per_cuz = 4
    schedule = []

    current_date = start_date

    while current_date <= end_date:
        sections_read = []
        for i, start_point in enumerate(start_points):
            current_cuz, current_hizb = start_point
            section_label = f"{current_cuz}. Cüz {current_hizb}. Hizb"
            sections_read.append(section_label)

            # Aktualisiere den Startpunkt für die nächste Person
            next_hizb = current_hizb + 1
            next_cuz = current_cuz
            if next_hizb > num_hizb_per_cuz:
                next_hizb = 1
                next_cuz += 1
                if next_cuz > num_cuz:
                    next_cuz = 1

            start_points[i] = (next_cuz, next_hizb)

        schedule.append([current_date.strftime("%d.%m.%Y")] + sections_read)

        current_date += timedelta(days=1)

    return schedule

def apply_column_colors(ws, people_names):
    for idx, column_name in enumerate(people_names, start=2):
        column_index = idx
        column_letter = chr(64 + column_index)  # Umwandlung in Spaltenbuchstabe (A, B, ...)
        random_color = ''.join([random.choice('0123456789ABCDEF') for _ in range(6)])  # Generiere eine zufällige Farbe
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=column_index, max_col=column_index):
            for cell in row:
                cell.fill = PatternFill(start_color=random_color, end_color=random_color, fill_type="solid")

def save_to_excel(schedule, people_names):
    # Filtere den Zeitplan auf nur einen Tag pro Woche
    filtered_schedule = []
    current_week = -1
    for row in schedule:
        date = datetime.strptime(row[0], "%d.%m.%Y")
        week_number = date.isocalendar()[1]

        if week_number != current_week:
            filtered_schedule.append(row)
            current_week = week_number

    wb = Workbook()
    ws = wb.active
    ws.append(['Datum'] + people_names)
    
    apply_column_colors(ws, people_names)
    
    for row in filtered_schedule:
        ws.append(row)
    
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f'reading_schedule_{current_time}.xlsx'
    wb.save(filename)

# Beispiel Einstellungen
start_date = datetime(2024, 4, 29)
end_date = datetime(2025, 2, 28)
people_names = ['personA', 'personB', 'personC']

schedule = generate_schedule(start_date, end_date, people_names)
save_to_excel(schedule, people_names)

